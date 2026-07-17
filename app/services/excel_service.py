"""Export des données au format Excel (OpenPyXL).

- ``export_report``   : rapport d'activité (3 feuilles : Synthèse,
  Ventes, Meilleurs produits).
- ``export_products`` : inventaire / catalogue produits.

Mise en forme : en-têtes colorées et figées, bordures fines, largeurs
automatiques, format monétaire sur les colonnes de montants.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Styles communs
# ------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="0A84FF")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1C1C1E")
_THIN = Side(style="thin", color="D1D1D6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row: int, ncols: int) -> None:
    """Applique le style d'en-tête sur une ligne."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _autosize(ws) -> None:
    """Ajuste la largeur des colonnes au contenu."""
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 4, 12), 50)


def _write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    """Écrit un tableau stylé, retourne la ligne suivante libre."""
    for j, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=header)
    _style_header_row(ws, start_row, len(headers))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    for i, row in enumerate(rows, start=start_row + 1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = _BODY_FONT
            cell.border = _BORDER
    return start_row + len(rows) + 1


# ------------------------------------------------------------------
# Rapport d'activité
# ------------------------------------------------------------------
def export_report(report: dict, dest: str | Path) -> str:
    """Exporte un rapport d'activité en classeur Excel.

    :param report: même structure que ``pdf_service.generate_report_pdf``.
    """
    dest = str(dest)
    wb = Workbook()

    # Feuille 1 : synthèse
    ws = wb.active
    ws.title = "Synthèse"
    ws.cell(row=1, column=1, value=report["title"]).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Période : {report['period_label']}").font = _BODY_FONT
    _write_table(ws, 4, ["Indicateur", "Valeur"], [list(r) for r in report["summary"]])
    _autosize(ws)

    # Feuille 2 : détail des ventes
    ws2 = wb.create_sheet("Ventes")
    _write_table(
        ws2, 1,
        ["N° facture", "Date", "Client", "Total", "Payé", "Reste"],
        [list(r) for r in report.get("sales", [])],
    )
    _autosize(ws2)

    # Feuille 3 : meilleurs produits
    ws3 = wb.create_sheet("Meilleurs produits")
    _write_table(
        ws3, 1,
        ["Produit", "Quantité", "Montant"],
        [list(r) for r in report.get("top_products", [])],
    )
    _autosize(ws3)

    wb.save(dest)
    return dest


# ------------------------------------------------------------------
# Catalogue / inventaire produits
# ------------------------------------------------------------------
def export_products(products: list[dict], dest: str | Path, title: str = "Inventaire des produits") -> str:
    """Exporte la liste des produits (catalogue ou inventaire)."""
    dest = str(dest)
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    rows = [
        [
            p["name"],
            p.get("sku", ""),
            p.get("category", ""),
            p.get("purchase_price", 0),
            p.get("sale_price", 0),
            p.get("stock_qty", 0),
            round(p.get("sale_price", 0) * p.get("stock_qty", 0), 2),
            "En stock" if p.get("stock_qty", 0) > 0 else "Rupture",
        ]
        for p in products
    ]
    _write_table(
        ws, 3,
        ["Produit", "Référence", "Catégorie", "Prix d'achat", "Prix de vente",
         "Stock", "Valeur du stock", "Statut"],
        rows,
    )
    _autosize(ws)
    wb.save(dest)
    return dest
